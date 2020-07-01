import openpyxl as op

# raw_input("Please enter the full file location. e.g. C:\Users\USER\Documents\Workbook.xlsx")
print("Currently, the address of the workbook will have to be the address of the code file")

# Extract file location and sheet
workbook_location_initial = input("What is the name of the excel file in this directory(e.g. classNames.xlsx)") #<<<< Enter the workbook location in this string <<<<
workbook_sheet = input("What is the name of the sheet with the data? e.g. Sheet1")
workbook = op.load_workbook((workbook_location_initial))
sheet = workbook[workbook_sheet]

# Ask which row has the first stock name

stockNameGO = 0
while stockNameGO == 0:
    try:
        name_row = int(input("In which row are the Stock Names placed? Please enter an integer etc."))
    except ValueError:
        print("Enter an Integer!")
        stockNameGO = 0
    else:
        stockNameGO = 1

# Ask which column has the first Stock name

name_column = input("From which column do the Stock Names start? Please enter A,B,C,D, etc.")
date_column = input("Which column contains the dates at which the stock prices were taken? Please enter A,B,C,D, etc.")
def column_to_cellnum(alphabet):
    order = 1
    letterGO = 0
    for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        if alphabet == letter:
            cellnum = order
        else:
            order = order + 1
    return(cellnum)

init_Row = name_row + 1 # Row in which first price is
init_Col = column_to_cellnum(name_column) # Stock's column
Row = init_Row # Changing Row and Column variables, initially set to first stock's column but next row from the name
Col = init_Col
position = sheet.cell(Row, Col).value

# MW - Main while loop that carries out iteration through stocks and respective price lists and displays anomolous outcomes

while position != None:
    PriceList = []
    stock_name = sheet.cell(Row - 1, Col).value
    # To check whether numbers are put in and not other characters
    stockPSGO = 0
    while stockPSGO == 0:
        try:
            percentage_significance = float(input("What percentage difference within dates would you consider significant for " + stock_name + "?"))
            percentage_significance_overall = float(input("What percentage difference from the initial date would you consider significant for " + stock_name + "?"))
        except ValueError:
            print("Enter an number!")
            stockPSGO = 0
        else:
            stockPSGO = 1
# MW, While loop that fills price list with each time's price
    while position != None:
        PriceList.append(position)
        Row = Row + 1
        position = sheet.cell(Row, Col).value
    print(stock_name + "'s prices are")
    print(PriceList)
# MW, For loop that iterates through the stock's price list and checks major jumps
    for value in range(len(PriceList)-1):
        n_0 = PriceList[value]
        n_1 = PriceList[value + 1]
        percentage_change = abs(((n_1-n_0)/(n_0))*100)
        percentage_change_non_abs = ((n_1-n_0)/(n_0))*100
        if (percentage_change > percentage_significance):
            print(stock_name + " has a significant change of " + percentage_change_non_abs + "% from " + str(sheet[date_column + str(value + init_Row)].value) + " to " + str(sheet[date_column + str((value + 1) + init_Row)].value))

    Row = init_Row # MW, set the row back to the initial 'second' row
    Col = Col + 1 # MW, iterate column to change stock
    position = sheet.cell(Row, Col).value










